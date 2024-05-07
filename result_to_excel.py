import datetime
import openpyxl
import os
from datetime import datetime
# Funkcja do zapisywania wyników w istniejącym pliku Excel
def save_results_to_excel(file_path, _in_,_out_, video_duration):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Odczytujemy istniejący plik Excel
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # Odczytujemy ostatni używany wiersz
    last_row = ws.max_row

    # Dodajemy nowy wiersz pod poprzednimi wynikami
    ws.cell(row=last_row + 1, column=1, value=timestamp)
    ws.cell(row=last_row + 1, column=2, value=_in_)
    ws.cell(row=last_row + 1, column=3, value=_out_)
    ws.cell(row=last_row + 1, column=4, value=video_duration)

    # Zapisujemy plik Excel
    wb.save(file_path)
    wb.save(file_path)

def if_excel_exist(file_path):
    if not os.path.exists(file_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Timestamp", "Out", "In", "Video Duration"])
        wb.save(file_path)